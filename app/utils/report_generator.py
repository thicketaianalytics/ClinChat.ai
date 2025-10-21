import reflex as rx
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Any

PRIMARY_COLOR = "#3b82f6"
PRIMARY_COLOR_LIGHT = "#dbeafe"
TEXT_COLOR = "#374151"
HEADER_TEXT_COLOR = "#FFFFFF"


def _add_header_footer(canvas, doc):
    """Adds a header and footer to each page of the PDF."""
    canvas.saveState()
    canvas.setFillColor(colors.HexColor(PRIMARY_COLOR))
    canvas.rect(
        doc.leftMargin,
        doc.height + doc.topMargin - 0.5 * inch,
        doc.width,
        0.4 * inch,
        fill=1,
        stroke=0,
    )
    canvas.setFont("Helvetica-Bold", 16)
    canvas.setFillColor(colors.white)
    canvas.drawString(
        doc.leftMargin + 0.2 * inch,
        doc.height + doc.topMargin - 0.35 * inch,
        "ClinChat.ai",
    )
    canvas.setFont("Helvetica", 9)
    canvas.setFillColor(colors.grey)
    canvas.line(
        doc.leftMargin,
        doc.bottomMargin - 0.2 * inch,
        doc.leftMargin + doc.width,
        doc.bottomMargin - 0.2 * inch,
    )
    canvas.drawString(
        doc.leftMargin,
        doc.bottomMargin - 0.4 * inch,
        f"Report Generated: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )
    canvas.drawRightString(
        doc.leftMargin + doc.width, doc.bottomMargin - 0.4 * inch, f"Page {doc.page}"
    )
    canvas.restoreState()


def _get_styles():
    """Returns a dictionary of ParagraphStyle objects."""
    styles = getSampleStyleSheet()
    return {
        "Title": ParagraphStyle(
            "Title",
            parent=styles["h1"],
            fontSize=22,
            textColor=PRIMARY_COLOR,
            spaceAfter=20,
            alignment=0,
        ),
        "Heading1": ParagraphStyle(
            "Heading1",
            parent=styles["h2"],
            fontSize=16,
            textColor=PRIMARY_COLOR,
            spaceBefore=12,
            spaceAfter=8,
            keepWithNext=1,
        ),
        "Heading2": ParagraphStyle(
            "Heading2",
            parent=styles["h3"],
            fontSize=12,
            textColor=TEXT_COLOR,
            spaceBefore=10,
            spaceAfter=6,
            keepWithNext=1,
        ),
        "Body": ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=10,
            textColor=TEXT_COLOR,
            leading=14,
            spaceAfter=6,
        ),
        "Bullet": ParagraphStyle(
            "Bullet",
            parent=styles["Normal"],
            fontSize=10,
            textColor=TEXT_COLOR,
            leftIndent=20,
            spaceAfter=4,
        ),
        "Code": ParagraphStyle(
            "Code",
            parent=styles["Code"],
            fontSize=9,
            textColor=colors.darkred,
            backColor=colors.ghostwhite,
            padding=4,
        ),
    }


def generate_trial_detail_pdf_from_state(
    state: "app.states.trial_detail_state.TrialDetailState",
) -> bytes | None:
    """Generates a PDF for a single trial from the TrialDetailState."""
    trial = state.trial
    if not trial:
        return None
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=1.0 * inch,
        bottomMargin=0.75 * inch,
    )
    story = []
    styles = _get_styles()
    story.append(Paragraph(trial.get("brief_title", "Trial Report"), styles["Title"]))
    story.append(
        Paragraph(f"<b>NCT ID:</b> {trial.get('nct_id', 'N/A')}", styles["Body"])
    )
    story.append(Spacer(1, 0.2 * inch))
    summary_data = [
        [
            Paragraph("<b>Status</b>", styles["Body"]),
            trial.get("overall_status", "N/A"),
        ],
        [Paragraph("<b>Phase</b>", styles["Body"]), trial.get("phase", "N/A")],
        [
            Paragraph("<b>Enrollment</b>", styles["Body"]),
            str(trial.get("enrollment", "N/A")),
        ],
        [
            Paragraph("<b>Start Date</b>", styles["Body"]),
            trial.get("start_date", "N/A"),
        ],
        [
            Paragraph("<b>Completion Date</b>", styles["Body"]),
            trial.get("completion_date", "N/A"),
        ],
        [
            Paragraph("<b>Complexity</b>", styles["Body"]),
            f"{state.complexity_rating} ({state.complexity_score} pts)",
        ],
    ]
    summary_table = Table(summary_data, colWidths=[1.5 * inch, 5.5 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(PRIMARY_COLOR_LIGHT)),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.2 * inch))
    if trial.get("brief_summary"):
        story.append(Paragraph("Brief Summary", styles["Heading1"]))
        story.append(Paragraph(trial["brief_summary"], styles["Body"]))
        story.append(Spacer(1, 0.1 * inch))
    if state.inclusion_criteria:
        story.append(Paragraph("Inclusion Criteria", styles["Heading1"]))
        for item in state.inclusion_criteria:
            story.append(Paragraph(f"• {item}", styles["Bullet"]))
        story.append(Spacer(1, 0.1 * inch))
    if state.exclusion_criteria:
        story.append(Paragraph("Exclusion Criteria", styles["Heading1"]))
        for item in state.exclusion_criteria:
            story.append(Paragraph(f"• {item}", styles["Bullet"]))
        story.append(Spacer(1, 0.1 * inch))
    doc.build(story, onFirstPage=_add_header_footer, onLaterPages=_add_header_footer)
    return buffer.getvalue()


def generate_comparison_pdf_from_state(
    state: "app.states.comparison_state.ComparisonState",
) -> bytes | None:
    """Generates a comparison PDF from the ComparisonState."""
    data = state.comparison_data
    if not data:
        return None
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=1.0 * inch,
        bottomMargin=0.75 * inch,
    )
    story = []
    styles = _get_styles()
    story.append(Paragraph("Clinical Trial Comparison Report", styles["Title"]))
    story.append(Spacer(1, 0.2 * inch))
    headers = ["Feature"] + [d.get("nct_id", "N/A") for d in data]
    rows = [
        ["Brief Title"]
        + [Paragraph(d.get("brief_title", "N/A"), styles["Body"]) for d in data],
        ["Status"] + [d.get("overall_status", "N/A") for d in data],
        ["Phase"] + [d.get("phase", "N/A") for d in data],
        ["Enrollment"] + [str(d.get("enrollment", "N/A")) for d in data],
        ["Start Date"] + [d.get("start_date", "N/A") for d in data],
        ["Completion Date"] + [d.get("completion_date", "N/A") for d in data],
        ["Study Type"] + [d.get("study_type", "N/A") for d in data],
    ]
    table_data = [headers] + rows
    table = Table(table_data, hAlign="LEFT", repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_COLOR)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor(PRIMARY_COLOR_LIGHT)),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ]
        )
    )
    story.append(table)
    doc.build(story, onFirstPage=_add_header_footer, onLaterPages=_add_header_footer)
    return buffer.getvalue()


def generate_excel_report_from_state(
    state: "app.states.saved_trials_state.SavedTrialsState",
) -> bytes | None:
    """Generates a multi-sheet Excel report from SavedTrialsState."""
    trials = state.saved_trials
    if not trials:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saved Trials Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color=PRIMARY_COLOR.replace("#", ""),
        end_color=PRIMARY_COLOR.replace("#", ""),
        fill_type="solid",
    )
    center_align = Alignment(horizontal="center", vertical="center")
    headers = [
        "NCT ID",
        "Brief Title",
        "Status",
        "Phase",
        "Enrollment",
        "Start Date",
        "Completion Date",
        "Tags",
        "Notes",
    ]
    ws.append(headers)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    for trial in trials:
        tags = ", ".join(trial.get("tags", []))
        ws.append(
            [
                trial.get("nct_id"),
                trial.get("brief_title"),
                trial.get("overall_status"),
                trial.get("phase"),
                trial.get("enrollment"),
                trial.get("start_date"),
                trial.get("completion_date"),
                tags,
                trial.get("notes"),
            ]
        )
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["I"].width = 40
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_analytics_pdf() -> bytes | None:
    return None